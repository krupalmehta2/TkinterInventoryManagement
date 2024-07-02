import tkinter as tk
from tkinter import messagebox
import openpyxl
import os

# Function to save or update inventory
def save_inventory():
    product_name = entry_product_name.get()
    brand = entry_brand.get()
    price = entry_price.get()
    qty = entry_qty.get()

    if not product_name or not brand or not price or not qty:
        messagebox.showerror("Input Error", "All fields are required!")
        return

    try:
        price = float(price)
        qty = int(qty)
    except ValueError:
        messagebox.showerror("Input Error", "Price must be a number and Quantity must be an integer!")
        return

    if not os.path.exists('inventory.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inventory"
        sheet.append(["Product Name", "Brand", "Price", "Quantity"])
    else:
        workbook = openpyxl.load_workbook('inventory.xlsx')
        sheet = workbook.active

    # Check if the product already exists
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == product_name:
            # Update the existing row
            row_idx = row[0].row
            sheet.cell(row=row_idx, column=2, value=brand)
            sheet.cell(row=row_idx, column=3, value=price)
            sheet.cell(row=row_idx, column=4, value=qty)
            workbook.save('inventory.xlsx')
            messagebox.showinfo("Success", "Product updated successfully!")
            return

    # Add new product
    sheet.append([product_name, brand, price, qty])
    workbook.save('inventory.xlsx')
    messagebox.showinfo("Success", "Product added successfully!")
    clear_entries()

def clear_entries():
    entry_product_name.delete(0, tk.END)
    entry_brand.delete(0, tk.END)
    entry_price.delete(0, tk.END)
    entry_qty.delete(0, tk.END)

# Create the main window
root = tk.Tk()
root.title("Inventory Management")

# Create and place labels and entries
tk.Label(root, text="Product Name").grid(row=0, column=0, padx=10, pady=10)
entry_product_name = tk.Entry(root)
entry_product_name.grid(row=0, column=1, padx=10, pady=10)

tk.Label(root, text="Brand").grid(row=1, column=0, padx=10, pady=10)
entry_brand = tk.Entry(root)
entry_brand.grid(row=1, column=1, padx=10, pady=10)

tk.Label(root, text="Price").grid(row=2, column=0, padx=10, pady=10)
entry_price = tk.Entry(root)
entry_price.grid(row=2, column=1, padx=10, pady=10)

tk.Label(root, text="Quantity").grid(row=3, column=0, padx=10, pady=10)
entry_qty = tk.Entry(root)
entry_qty.grid(row=3, column=1, padx=10, pady=10)

# Create and place buttons
tk.Button(root, text="Add/Update Product", command=save_inventory).grid(row=4, column=0, columnspan=2, pady=10)
tk.Button(root, text="Clear", command=clear_entries).grid(row=5, column=0, columnspan=2, pady=10)

# Run the application
root.mainloop()
